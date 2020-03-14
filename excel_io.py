#!/usr/bin/env python3

import openpyxl

MAX_STUDENTS = 28

###
## Read class list from Excel file
###
def read_class_lists(infile):
    wb = openpyxl.load_workbook(filename = infile)
    sheet = wb.active
    lis = []
    for i in range(sheet.max_row - 1):
        lis.append(sheet.cell(row = i + 2, column = 1).value)
    return lis
